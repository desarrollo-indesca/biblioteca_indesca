from django.core.management.base import BaseCommand
from django.db import transaction
from core.models import Informe, Descriptor, Empresa
from biblioteca_digital.settings import BASE_DIR
import openpyxl

class Command(BaseCommand):
    help = "Carga los informes en la base de datos junto a los posibles descriptores"

    def handle(self, *args, **options):
        dataframe = openpyxl.load_workbook(BASE_DIR.__str__() + "\\INFORMES.xlsx")['INFORMES']
        dataframe.active = True
        with transaction.atomic():
            for row in dataframe.iter_rows(2, dataframe.max_row):
                informe = Informe()

                no_registro = row[0].value
                if not no_registro or no_registro == '':
                    continue

                ano = row[9].value
                ano = ano if ano != '' else None
                if not ano or ano == '':
                    continue

                if Informe.objects.filter(no_registro=no_registro).exists():
                    continue

                autores = row[4].value
                if not autores or autores == '':
                    continue

                programa = str(row[1].value).strip().upper()
                if not programa or programa == '' or programa == 'NONE':
                    continue
                
                descriptor = str(row[8].value).strip().upper()
                solicitud_servicio = str(row[5].value).strip().upper().replace("_X000D_", "").strip() if row[5].value else None
                informe.no_registro = no_registro.strip()
                informe.programa = Empresa.objects.get_or_create(nombre=programa)[0]
                informe.codigo_proyecto = str(row[2].value).strip()
                informe.ano_publicacion = ano
                informe.titulo = str(row[7].value).strip().upper().replace("\"", "").replace("_X000D_", "").strip()
                informe.autores = autores.strip().upper().replace("_X000D_", "")
                informe.solicitud_servicio = solicitud_servicio if solicitud_servicio != '' and solicitud_servicio != 'None' else None
                archivo = "/".join(row[-2].value.split('/')[-3:]) if row[-2].value and len(row[-2].value) and row[-2].value != 'documentone.htm' and  "/".join(row[-2].value.split('/')[-3:]) != '' else None

                if(archivo):
                    archivo = archivo.replace("\INFORMES\\", "").replace("\informes\\",'').replace("informes\\",'')
                    informe.archivo = archivo
                else:
                    continue

                print(f"{informe.archivo}")

                informe.save()

                descriptores = []                    
                for x in descriptor.split(','):
                    x = x.strip().upper()
                    if(x == '' or x == 'NONE'):
                        continue
                    descriptor = Descriptor.objects.get_or_create(nombre=x)[0]
                    descriptores.append(descriptor)
                
                informe.descriptores.add(*descriptores)
                informe.save()